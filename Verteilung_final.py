from csv import writer
from csv import reader
import csv
import pandas
from IPython.display import display
pandas.options.display.max_columns=30
pandas.options.display.max_rows=30
from operator import add
import numpy
import xlsxwriter
import openpyxl

####
#Data Wrangling
wb = openpyxl.load_workbook("Anmeldungen_test.xlsx")

sheet = wb["Tabelle1"]

#Unnötige Daten aus dem Formular werden gelöscht 
sheet.delete_cols(1,3)
sheet.delete_cols(15,1)
sheet.insert_cols(10,1)
sheet.insert_cols(13,1)
sheet.insert_cols(6,1)

#Benötigte Spalten werden geschrieben
sheet.cell(row=1, column=6).value = "Kriterium"
sheet.cell(row=1, column=11).value = "Überschneidung HK"
sheet.cell(row=1, column=14).value = "Überschneidung NK"

sheet.cell(row=1, column=1).value = "Prüfungsname"
sheet.cell(row=1, column=2).value = "LV"
sheet.cell(row=1, column=3).value = "FB"
sheet.cell(row=1, column=4).value = "Studierende"
sheet.cell(row=1, column=5).value = "Dauer"
sheet.cell(row=1, column=7).value = "Format"
sheet.cell(row=1, column=8).value = "Erfahrung"
sheet.cell(row=1, column=9).value = "Datum_HK"
sheet.cell(row=1, column=10).value = "Zeit_HK"
sheet.cell(row=1, column=12).value = "Datum_NK"
sheet.cell(row=1, column=13).value = "Zeit_NK"
sheet.cell(row=1, column=13).value = "Zeit_NK"
sheet.cell(row=1, column=15).value = "Namen"
sheet.cell(row=1, column=16).value = "Anmerkung"
sheet.cell(row=1, column=17).value = "Mail"

#Excel wird gespeichert
wb.save("Anmeldungen_test.xlsx")

#####Ende Data Wrangling#####

#Datensatz wird in Dataframe eingelesen
df=pandas.read_excel("Anmeldungen_test.xlsx")


#Umgewandelte Datensätze werden in dt. Format ungewandelt
df["Datum_HK"]=df["Datum_HK"].dt.strftime("%d.%m.%Y")
df["Datum_NK"]=df["Datum_NK"].dt.strftime("%d.%m.%Y")

#Datei wird als CSV gespeichert, um sie mit dem csv Modul einlesen zu können.
df.to_csv('Anmeldungen_test.csv', index=False, sep=";", encoding='iso-8859-1')

#Variable wird initiert, um Position im CSV Dokument zu bestimmen
x=-1

#Einlesen des Datensatzes
df = pandas.read_csv('Anmeldungen_test.csv',encoding = 'iso-8859-1', sep=';')

#Es wird eine Masterliste mit den gemeinsamen Daten und Uhrzeiten der HK generiert
liste1=df["Datum_HK"].values.tolist()
liste2=df["Zeit_HK"].values.tolist()

z=zip(liste1,liste2)

e=list(z)

#Es wird eine Masterliste mit den gemeinsamen Daten und Uhrzeiten der NK generiert
liste_datum_nk=df["Datum_NK"].values.tolist()
liste_uhrzeit_nk=df["Zeit_NK"].values.tolist()

zip_nk_gemeinsam=zip(liste_datum_nk,liste_uhrzeit_nk)

liste_nk_gemeinsam=list(zip_nk_gemeinsam)

#CSV wird geöffnet und es wird über alle Reihen iteriert
with open ('Anmeldungen_test.csv') as csv_file:
    csv_reader=csv.DictReader(csv_file,delimiter=';')
    line_count=0
    for row in csv_reader:
                #Nummer=row['Nummer']
                Veranstaltung=row['Prüfungsname']
                LVNummer=row['LV']
                Fachbereich=row['FB']
                Anzahl_Studierende=row['Studierende']
                Prüfungsdauer=row['Dauer']
                DistanzOPräsenz=row['Format']
                Datum_HK=row['Datum_HK']
                Uhrzeit_HK=row['Zeit_HK']
                Datum_NK=row['Datum_NK']
                Uhrzeit_NK=row['Zeit_NK']
                Lehrende=row['Namen']
                Mail=row['Mail']
                Kriterium=row['Kriterium']
                Überschneidung_HK=row['Überschneidung HK']
                Überschneidung_NK=row['Überschneidung NK']
                Erfahrung=row['Erfahrung']
                Anmerkung=row["Anmerkung"]

                #Variable zählt hoch, um Zeile zu bestimmen
                x=x+1

                #Falsche Formatierung der Daten wird geändert
                if Anzahl_Studierende == "1940-01-01 00:00:00":
                    df.at[x,"Studierende"] = "1-40P"  
                

                #Kriteriumsdefinitionen

                if Anzahl_Studierende == "mehr als 340" and Kriterium != "1":
                    df.at[x,"Kriterium"] = '2'

                if Anzahl_Studierende == "171-340" and Kriterium != "1":
                    df.at[x,"Kriterium"] = '3'

                if Anzahl_Studierende == "121-170" and Kriterium != "1":
                    df.at[x,"Kriterium"] = '4'


                if Anzahl_Studierende == "81-120" and Kriterium != "1":
                    df.at[x,"Kriterium"] = '5'

                if Anzahl_Studierende == "41-80" and Kriterium != "1":
                    df.at[x,"Kriterium"] = '6'

                
                if Anzahl_Studierende == "1-40P" and Kriterium != "1":
                    df.at[x,"Kriterium"] = '7'


                #Verteilung hoher Priorität für wichtige FBs

                if Fachbereich == "FB Veterinärmedizin":
                    df.at[x,"Kriterium"] = '1'

                if Fachbereich == "FB Biologie, Chemie, Pharmazie":
                    df.at[x,"Kriterium"] = '1'
                            
                if Fachbereich == "FB Rechtswissenschaft":
                    df.at[x,"Kriterium"] = "1"

                #Bearbeitung der HK Verteilung

                #Es wird eine Liste mit dem individuellen Datum der jeweiligen Prüfung angelegt
                liste3=[Datum_HK]
                liste4=[Uhrzeit_HK]
                
                q=zip(liste3,liste4)

                m= list(q)

                #Das zu analysierende Datum wird aus der Gesamtdatenliste genommen, um keinen false positive zu erhalten
                e.pop(0)

                #Schleife zur Überprüfung, ob sich das Datum überschneidet
                s=-1
                for i in e:
                    s=s+1

                    if m[0]==(e[s]):
 
                        o=True
                        break
                    
                    else:
                        o=False
                    
                    
                #Schreiben in DF, je nach Ergebnis der Überschneidungsanalyse
                if o == True:
                    
                    df = pandas.DataFrame(df)
                    df.at[x,"Überschneidung HK"] = 'Ja'
                
                
                elif o == False:
                    df = pandas.DataFrame(df)
                    df.at[x,"Überschneidung HK"] = 'Nein'

                #Das herausgenommene Datum wird wieder an die Masterliste angehangen
                list_10 = []
                list_10.extend(e + m)
                e=list_10



                #Bearbeitung der NK Verteilung

                #Es wird eine Liste mit dem individuellen Datum der jeweiligen Prüfung angelegt
                liste11=[Datum_NK]
                liste12=[Uhrzeit_NK]
                
                zip_aktuelles_datum_nk=zip(liste11,liste12)

                liste_aktuelles_datum_nk = list(zip_aktuelles_datum_nk)
                
                #Das zu analysierende Datum wird aus der Gesamtdatenliste genommen, um keinen false positive zu erhalten
                liste_nk_gemeinsam.pop(0) 

                #Schleife zur Überprüfung, ob sich das Datum überschneidet
                l=-1
                for i in liste_nk_gemeinsam:
                    l=l+1

                    if liste_aktuelles_datum_nk[0]==(liste_nk_gemeinsam[l]):

                        c=True
                        break
                    
                    else:
                        c=False
                    
                    
                #Schreiben in DF, je nach Ergebnis der Überschneidungsanalyse
                if c == True:

                    df = pandas.DataFrame(df)
                    df.at[x,"Überschneidung NK"] = 'Ja'
                
                
                elif c == False:

                    df = pandas.DataFrame(df)
                    df.at[x,"Überschneidung NK"] = 'Nein'

                #Das herausgenommene Datum wird wieder an die Masterliste angehangen
                list_13 = []
                list_13.extend(liste_nk_gemeinsam + liste_aktuelles_datum_nk)
                liste_nk_gemeinsam=list_13


#CSV wird geschrieben
df.to_csv('Anmeldungen_test.csv', index=False, sep=";", encoding='iso-8859-1')


###### Output der beiden individuellen Excel Dokumente für die HK Überschneidungen

#Dataframe wird generiert und Spalten werden definiert
df1=pandas.DataFrame(columns=["Prüfungsname",	"LV","FB","Studierende","Dauer","Format","Kriterium","Erfahrung","Datum_HK","Zeit_HK","Überschneidung HK","Datum_NK","Zeit_NK",
        "Überschneidung NK","Namen","Anmerkung","Mail"])

#Variable zur Verortung der zu beschreibenen Zelle
v=-1

#Über den Master-DF iterieren, um bei fehlender Überschneidung diese Row an DF1 zu hängen
for index, row in df.iterrows():
    v=v+1
    if df.at[v, 'Überschneidung HK'] == "Nein":

        df1=pandas.concat([df1, df.iloc[[v]]])

#Unnötige Spalten im DF1 löschen
del df1["Datum_NK"]
del df1["Zeit_NK"]
del df1["Überschneidung NK"]

#Datumswerte in tatsächliche Daten umwandeln, um sie sortieren zu können
df1["Datum_HK"] = pandas.to_datetime(df1["Datum_HK"], format="%d.%m.%Y")

#Sortieren der Rows nach Datum
df1=df1.sort_values(by=["Datum_HK"])

#Schreiben der HK Liste ohne Überschneidung
df1.to_excel('Überschneidungen_HK_Nein.xlsx', index=False, encoding='iso-8859-1')

#Öffnen der selben Datei, um Tabellenblatt umzubennen und die Daten richtig zu formatieren (Danke Excel!)
wb = openpyxl.load_workbook("Überschneidungen_HK_Nein.xlsx")
wb_sheet = wb['Sheet1']
wb_sheet.title = 'Keine Überschneidungen HK'

r=0
for row in wb["Keine Überschneidungen HK"]:
    r=r+1
    wb["Keine Überschneidungen HK"][str("I")+str(r)].number_format = 'dd/mm/yy'

wb.save("Überschneidungen_HK_Nein.xlsx")


#Selbe wie oben, nur für die NK
df2=pandas.DataFrame(columns=["Prüfungsname","LV","FB","Studierende","Dauer","Format","Kriterium","Erfahrung","Datum_HK","Zeit_HK","Überschneidung HK","Datum_NK","Zeit_NK",
        "Überschneidung NK","Namen","Anmerkung","Mail"])

j=-1
for index, row in df.iterrows():
    j=j+1
    if df.at[j, 'Überschneidung HK'] == "Ja":

        df2=pandas.concat([df2, df.iloc[[j]]])

del df2["Datum_NK"]
del df2["Zeit_NK"]
del df2["Überschneidung NK"]

df2["Datum_HK"] = pandas.to_datetime(df2["Datum_HK"], format="%d.%m.%Y")
df2=df2.sort_values(by=["Datum_HK"])
#df2.to_csv('Überschneidungen_HK_Ja.csv', index=False, sep=";", encoding='iso-8859-1')
df2.to_excel("Überschneidungen_HK_Ja.xlsx", index=False, encoding='iso-8859-1')


wb = openpyxl.load_workbook("Überschneidungen_HK_Ja.xlsx")
wb_sheet = wb['Sheet1']
wb_sheet.title = 'Überschneidungen HK'

r=0
for row in wb["Überschneidungen HK"]:
    r=r+1
    wb["Überschneidungen HK"][str("I")+str(r)].number_format = 'dd/mm/yy'
wb.save("Überschneidungen_HK_Ja.xlsx")

#### Output für die beiden Excel Dokumente für die NK

#Alles das Gleiche wie oben, nur für die NK
df3=pandas.DataFrame(columns=["Prüfungsname",	"LV","FB","Studierende","Dauer","Format","Kriterium","Erfahrung","Datum_HK","Zeit_HK","Überschneidung HK","Datum_NK","Zeit_NK",
        "Überschneidung NK","Namen","Anmerkung","Mail"])

v=-1
for index, row in df.iterrows():
    v=v+1
    if df.at[v, 'Überschneidung NK'] == "Nein":

        df3=pandas.concat([df3, df.iloc[[v]]])

del df3["Datum_HK"]
del df3["Zeit_HK"]
del df3["Überschneidung HK"]

df3["Datum_NK"] = pandas.to_datetime(df3["Datum_NK"], format="%d.%m.%Y")
df3=df3.sort_values(by=["Datum_NK"])

df3.to_excel('Überschneidungen_NK_Nein.xlsx', index=False, encoding='iso-8859-1')


wb = openpyxl.load_workbook("Überschneidungen_NK_Nein.xlsx")
wb_sheet = wb['Sheet1']
wb_sheet.title = 'Keine Überschneidungen NK'

r=0
for row in wb["Keine Überschneidungen NK"]:
    r=r+1
    wb["Keine Überschneidungen NK"][str("I")+str(r)].number_format = 'dd/mm/yy'

wb.save("Überschneidungen_NK_Nein.xlsx")



df4=pandas.DataFrame(columns=["Prüfungsname",	"LV","FB","Studierende","Dauer","Format","Kriterium","Erfahrung","Datum_HK","Zeit_HK","Überschneidung HK","Datum_NK","Zeit_NK",
        "Überschneidung NK","Namen","Anmerkung","Mail"])
v=-1
for index, row in df.iterrows():
    v=v+1
    if df.at[v, 'Überschneidung NK'] == "Ja":

        df4=pandas.concat([df4, df.iloc[[v]]])

del df4["Datum_HK"]
del df4["Zeit_HK"]
del df4["Überschneidung HK"]

df4["Datum_NK"] = pandas.to_datetime(df4["Datum_NK"], format="%d.%m.%Y")
df4=df4.sort_values(by=["Datum_NK"])

df4.to_excel("Überschneidungen_NK_Ja.xlsx", index=False, encoding='iso-8859-1')


wb = openpyxl.load_workbook("Überschneidungen_NK_Ja.xlsx")
wb_sheet = wb['Sheet1']
wb_sheet.title = 'Überschneidungen NK'

r=0
for row in wb["Überschneidungen NK"]:
    r=r+1
    wb["Überschneidungen NK"][str("I")+str(r)].number_format = 'dd/mm/yy'
wb.save("Überschneidungen_NK_Ja.xlsx")

####ENDE####
